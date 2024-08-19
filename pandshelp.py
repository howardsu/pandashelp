import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
def append_df_to_excel(df, filename, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    """
    将DataFrame追加到Excel文件中的指定工作表，包括表头。
    
    :param df: 要追加的DataFrame
    :param filename: 目标Excel文件的路径
    :param sheet_name: 要追加数据的工作表名称
    :param startrow: 开始写入数据的行号（如果为None，则从最后一行开始）
    :param truncate_sheet: 是否在写入前截断工作表
    :param to_excel_kwargs: 传递给to_excel方法的其他参数
    """
    
    # 如果文件不存在，直接写入新文件（包括表头）
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        df.to_excel(filename, sheet_name=sheet_name, **to_excel_kwargs)
        return
    
    # 打开目标工作表，如果不存在则创建
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        startrow = 0  # 新工作表从第一行开始写
    else:
        ws = wb[sheet_name]
    
    if truncate_sheet and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
        startrow = 0  # 如果截断了工作表，从第一行开始写
    
    if startrow is None:
        startrow = ws.max_row
    
    # 总是写入表头，除非追加到非空的现有工作表
    write_header = startrow == 0
    
    # 写入数据（包括表头，如果需要的话）
    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        if not write_header and r_idx == 1:
            continue  # 如果不需要写入表头，跳过第一行
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=startrow + r_idx, column=c_idx, value=value)
    
    wb.save(filename)
